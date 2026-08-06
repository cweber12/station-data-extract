"""
exporter.py -- writes the generated comparison workbook.

Everything on the `stats` sheet is a live formula pointing at `data`, so if you
paste in a corrected series or trim a date range by hand, the numbers follow.
Only the lag table is precomputed, because scanning lags in Excel needs an
OFFSET mess that would be worse than the thing it replaces.
"""

from __future__ import annotations

import re

from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd

import identity
import sensorkit as sk
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, ScatterChart, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import (CharacterProperties, Font as Font_,
                                   Paragraph, ParagraphProperties,
                                   RichTextProperties)
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

LOCAL_TZ = ZoneInfo("America/Los_Angeles")

BODY = Font(name="Arial", size=10)
HEAD = Font(name="Arial", size=10, bold=True)
TITLE = Font(name="Arial", size=12, bold=True)
NOTE = Font(name="Arial", size=9, italic=True, color="808080")
HEADFILL = PatternFill("solid", fgColor="D9D9D9")
WARNFILL = PatternFill("solid", fgColor="FFF2CC")

DT_FMT = "yyyy-mm-dd hh:mm"
NUM_FMT = "0.000"


def _style_header(ws, row=1, ncols=None):
    ncols = ncols or ws.max_column
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEAD
        cell.fill = HEADFILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = ws.cell(row=row + 1, column=3)


def _autosize(ws, minw=10, maxw=34):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = max((len(str(c.value)) for c in col[:60] if c.value is not None),
                      default=0)
        ws.column_dimensions[letter].width = min(max(minw, longest + 2), maxw)


def _write_data_sheet(wb, result):
    ws = wb.create_sheet("data")
    cols = list(result.data.columns)

    ws.cell(row=1, column=1, value="time (local)")
    ws.cell(row=1, column=2, value="time (UTC)")
    for j, c in enumerate(cols, start=3):
        unit = result.units.get(c, "")
        ws.cell(row=1, column=j, value=f"{c} [{unit}]" if unit else c)

    local = result.data.index.tz_convert(LOCAL_TZ)
    for i, (tl, tu) in enumerate(zip(local, result.data.index), start=2):
        ws.cell(row=i, column=1, value=tl.replace(tzinfo=None)).number_format = DT_FMT
        ws.cell(row=i, column=2, value=tu.replace(tzinfo=None)).number_format = DT_FMT
        for j, c in enumerate(cols, start=3):
            v = result.data.iloc[i - 2][c]
            cell = ws.cell(row=i, column=j,
                           value=None if pd.isna(v) else float(v))
            cell.number_format = NUM_FMT

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY
    _style_header(ws)
    _autosize(ws)
    return ws, cols


def _write_counts_sheet(wb, result, cols):
    ws = wb.create_sheet("counts")
    ws.cell(row=1, column=1, value="time (local)")
    for j, c in enumerate(cols, start=2):
        ws.cell(row=1, column=j, value=c)

    local = result.data.index.tz_convert(LOCAL_TZ)
    expected = {}
    for c in cols:
        try:
            cad = pd.Timedelta(result.cadences.get(c, "NaT"))
            expected[c] = max(1, int(pd.Timedelta(result.interval) / cad))
        except Exception:
            expected[c] = 1

    for i, tl in enumerate(local, start=2):
        ws.cell(row=i, column=1, value=tl.replace(tzinfo=None)).number_format = DT_FMT
        for j, c in enumerate(cols, start=2):
            n = int(result.counts.iloc[i - 2][c]) if c in result.counts else 0
            cell = ws.cell(row=i, column=j, value=n)
            if n < expected[c] / 2:
                cell.fill = WARNFILL

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY
    _style_header(ws)
    _autosize(ws)

    r = ws.max_row + 2
    ws.cell(row=r, column=1,
            value="Shaded = fewer than half the samples expected for this "
                  "interval, given the series' native cadence. Treat those "
                  "averages as thin.").font = NOTE
    return ws


def _write_stats_sheet(wb, result, cols, lag_table=None, reference=None):
    ws = wb.create_sheet("stats")
    n = len(result.data)
    last = n + 1  # data sheet last row

    def dref(c_index):
        L = get_column_letter(c_index + 3)
        return f"data!${L}$2:${L}${last}"

    ws.cell(row=1, column=1, value="Summary").font = TITLE
    heads = ["series", "unit", "n", "mean", "std", "min", "max", "native cadence"]
    for j, h in enumerate(heads, start=1):
        ws.cell(row=2, column=j, value=h)
    for i, c in enumerate(cols):
        r = 3 + i
        ws.cell(row=r, column=1, value=c)
        ws.cell(row=r, column=2, value=result.units.get(c, ""))
        ws.cell(row=r, column=3, value=f"=COUNT({dref(i)})")
        ws.cell(row=r, column=4, value=f"=AVERAGE({dref(i)})").number_format = NUM_FMT
        ws.cell(row=r, column=5, value=f"=STDEV({dref(i)})").number_format = NUM_FMT
        ws.cell(row=r, column=6, value=f"=MIN({dref(i)})").number_format = NUM_FMT
        ws.cell(row=r, column=7, value=f"=MAX({dref(i)})").number_format = NUM_FMT
        ws.cell(row=r, column=8, value=result.cadences.get(c, ""))
    _style_header(ws, row=2, ncols=len(heads))

    base = 3 + len(cols) + 2
    ws.cell(row=base, column=1, value="Correlation matrix (lag 0)").font = TITLE
    hr = base + 1
    for j, c in enumerate(cols, start=2):
        ws.cell(row=hr, column=j, value=c).font = HEAD
    for i, c in enumerate(cols):
        r = hr + 1 + i
        ws.cell(row=r, column=1, value=c).font = HEAD
        for j, _ in enumerate(cols):
            cell = ws.cell(row=r, column=2 + j)
            if i == j:
                cell.value = 1.0
            else:
                cell.value = f"=IFERROR(CORREL({dref(i)},{dref(j)}),\"\")"
            cell.number_format = "0.000"

    r = hr + len(cols) + 3
    if lag_table is not None and len(lag_table):
        ws.cell(row=r, column=1,
                value=f"Lag scan against: {reference}").font = TITLE
        r += 1
        lheads = ["series", "r at lag 0", "lags reference by (h)",
                  "r at best lag", "alt lag (h), 12.42 h ambiguity",
                  "n overlapping"]
        for j, h in enumerate(lheads, start=1):
            ws.cell(row=r, column=j, value=h)
        _style_header(ws, row=r, ncols=len(lheads))
        ws.freeze_panes = None
        for _, rec in lag_table.iterrows():
            r += 1
            ws.cell(row=r, column=1, value=rec["series"])
            ws.cell(row=r, column=2, value=float(rec["r_at_lag_0"])).number_format = "0.000"
            ws.cell(row=r, column=3, value=float(rec["best_lag_h"])).number_format = "0.00"
            ws.cell(row=r, column=4, value=float(rec["r_at_best_lag"])).number_format = "0.000"
            ws.cell(row=r, column=5, value=float(rec["ambiguous_alt_h"])).number_format = "0.00"
            ws.cell(row=r, column=6, value=int(rec["n_overlapping"]))
        r += 2
        ws.cell(row=r, column=1,
                value="Negative lag means the series leads the reference. "
                      "La Jolla nearshore temperature is dominated by the "
                      "~12.42 h internal tide, so a peak at lag L cannot be "
                      "distinguished from one at L +/- 12.42 h. Both are "
                      "listed. Resolve with physics, not with this table.").font = NOTE

    for row in ws.iter_rows():
        for cell in row:
            if cell.font is BODY or cell.font.name is None:
                cell.font = BODY
    _autosize(ws)
    return ws


def _write_zscore_sheet(wb, result, cols):
    """Standardised copy, so series in different units share one axis."""
    ws = wb.create_sheet("normalized")
    z = (result.data - result.data.mean()) / result.data.std(ddof=0)

    ws.cell(row=1, column=1, value="time (local)")
    for j, c in enumerate(cols, start=2):
        ws.cell(row=1, column=j, value=c)

    local = result.data.index.tz_convert(LOCAL_TZ)
    for i, tl in enumerate(local, start=2):
        ws.cell(row=i, column=1, value=tl.replace(tzinfo=None)).number_format = DT_FMT
        for j, c in enumerate(cols, start=2):
            v = z.iloc[i - 2][c]
            ws.cell(row=i, column=j,
                    value=None if pd.isna(v) else float(v)).number_format = NUM_FMT

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY
    _style_header(ws)
    _autosize(ws)

    r = ws.max_row + 2
    ws.cell(row=r, column=1,
            value="Each series minus its own mean, divided by its own standard "
                  "deviation. Use this chart to compare shape and timing across "
                  "different units; use the raw chart for magnitudes.").font = NOTE
    return ws


GRID = "D9D9D9"
AXIS = "808080"
EXCEL_EPOCH = datetime(1899, 12, 30)

# Nice tick spacings in days, smallest first.
TICK_STEPS = [1 / 24, 2 / 24, 3 / 24, 6 / 24, 12 / 24, 1, 2, 3, 5, 7, 14, 30]


def _serial(dt) -> float:
    """Excel date serial for a naive datetime."""
    return (dt - EXCEL_EPOCH).total_seconds() / 86400.0


def _tick_step(span_days: float, target: int = 11) -> float:
    ideal = span_days / max(target, 1)
    for step in TICK_STEPS:
        if step >= ideal:
            return step
    return TICK_STEPS[-1]


def _text_props(size=900, rotation=None, bold=False):
    body = RichTextProperties(vert="horz", anchor="ctr", anchorCtr=True)
    if rotation is not None:
        body.rot = rotation
    chars = CharacterProperties(sz=size, b=bold,
                                latin=Font_(typeface="Arial"))
    return RichText(bodyPr=body,
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=chars),
                                 endParaRPr=chars)])


def _set_axis_title(axis, text, rotation=None, pin_left=False,
                    chart_h=None):
    """
    Set an axis title and force its rotation.

    Excel auto-places axis titles, and in a narrow chart it will happily drop a
    rotated title on top of the tick labels. `pin_left` overrides that with an
    explicit layout: hard against the left edge, vertically centred on the plot
    band, leaving the tick labels the space to its right.
    """
    axis.title = text
    if pin_left:
        px = (chart_h or CHART_H) / 2.54 * 96
        h_frac = min(0.6, (len(str(text)) * 5.5 + 10) / px)
        axis.title.layout = Layout(manualLayout=ManualLayout(
            xMode="edge", yMode="edge",
            x=0.0, y=round(PLOT_Y + PLOT_H / 2 - h_frac / 2, 4)))
    rich = axis.title.tx.rich
    body = RichTextProperties(vert="horz", anchor="ctr", anchorCtr=True)
    if rotation is not None:
        body.rot = rotation
    rich.bodyPr = body
    chars = CharacterProperties(sz=1000, b=False, latin=Font_(typeface="Arial"))
    for para in rich.p:
        para.pPr = ParagraphProperties(defRPr=chars)
        for run in (para.r or []):
            run.rPr = chars
    axis.title.overlay = False


def _style_common(ch):
    ch.style = None
    ch.dispBlanksAs = "gap"
    for ax in (ch.x_axis, ch.y_axis):
        ax.delete = False
        ax.majorTickMark = "out"
        ax.minorTickMark = "none"
        # "autoZero" would draw each axis where the other hits zero, which
        # lands mid-plot the moment a series goes negative (z-scores). Pin
        # both to the low end so the frame stays around the data.
        ax.crosses = "min"
        ax.tickLblPos = "low"
        ax.spPr = GraphicalProperties(
            ln=LineProperties(solidFill=AXIS, w=9525))
        ax.txPr = _text_props(900)
    gl = ChartLines()
    gl.spPr = GraphicalProperties(ln=LineProperties(solidFill=GRID, w=9525))
    ch.y_axis.majorGridlines = gl
    ch.x_axis.majorGridlines = None
    if ch.legend is not None:
        ch.legend.position = "b"
        ch.legend.overlay = False
        ch.legend.txPr = _text_props(900)


def _style_time_axis(ch, index, title="time (local)", target_ticks=11):
    """A real numeric axis carrying date serials -- no category collapsing."""
    lo = _serial(index[0].tz_convert(LOCAL_TZ).replace(tzinfo=None))
    hi = _serial(index[-1].tz_convert(LOCAL_TZ).replace(tzinfo=None))
    step = _tick_step(hi - lo, target_ticks)
    ch.x_axis.scaling.min = round(lo, 6)
    ch.x_axis.scaling.max = round(hi, 6)
    ch.x_axis.majorUnit = step
    ch.x_axis.number_format = "mm-dd hh:mm" if step < 1 else "mm-dd"
    ch.x_axis.txPr = _text_props(900, rotation=-2700000)
    _set_axis_title(ch.x_axis, title)


def _style_value_axis(ch, frame, cols, title, pin_left=False):
    vals = frame[cols].to_numpy(dtype="float64")
    vals = vals[~np.isnan(vals)]
    if vals.size:
        lo, hi = float(vals.min()), float(vals.max())
        if hi == lo:
            lo, hi = lo - 1, hi + 1
        pad = (hi - lo) * 0.08
        ch.y_axis.scaling.min = round(lo - pad, 2)
        ch.y_axis.scaling.max = round(hi + pad, 2)
    if title:
        _set_axis_title(ch.y_axis, title, rotation=-5400000,
                        pin_left=pin_left)


def _add_line_series(ch, ws, col_ix, nrows, hexes, smooth=False):
    """One straight-line series per column, no markers.

    `col_ix` is the list of sheet column indices to plot, in legend order. It
    is a list rather than a start-plus-count because a chart carries one unit
    group, and a group is an arbitrary subset of the sheet's columns --
    salinity can sit between two temperatures.

    `hexes` is the matching list of colours. Passed in rather than taken from
    position, because position is per-chart and a series has to keep one
    colour across every sheet it appears on.
    """
    xref = Reference(ws, min_col=1, min_row=2, max_row=nrows + 1)
    for i, c in enumerate(col_ix):
        yref = Reference(ws, min_col=c, min_row=1, max_row=nrows + 1)
        ser = Series(yref, xref, title_from_data=True)
        ser.smooth = smooth
        ser.marker = Marker(symbol="none")
        ser.graphicalProperties = GraphicalProperties(
            ln=LineProperties(solidFill=hexes[i], w=17000, cap="rnd"))
        ch.series.append(ser)


# Horizontal scale. Width is a function of ELAPSED TIME, not of how many points
# happen to land in it, so one window comes out one width whatever interval is
# chosen.
#
# It used to be points x 0.15 cm, which made the width a function of the
# interval: over 45 days a 3 h build was 54 cm and a 30 min build was 324 cm.
# Putting those side by side compares magnifications, not curves, so the
# question "which interval reads best here" could not actually be asked.
#
# 3.6 cm/day is what a 1 h build produced under the old rule (1,080 points x
# 0.15 cm across 45 days), so 1 h charts keep the scale they had and every
# other interval joins them on it.
#
# MIN_W only kicks in for very short records; MAX_W keeps Excel from choking on
# an absurd object. Either one binding breaks the constant-scale promise, so
# when one does chart_scale_note() reports it in the log, the dialog and the
# provenance sheet -- a workbook must never claim a scale it was not drawn at.
CM_PER_DAY = 3.6
MIN_W, MAX_W = 30.0, 800.0

AXIS_COL_W = 15.0                    # frozen column -- wide enough for the
                                     # rotated title AND the tick labels beside
                                     # it, with a margin so neither collides
SWATCH_W = 2.2                       # legend swatch columns
DEFAULT_COL_PX = 64                  # an untouched column, in pixels
CHAR_PX = 5.4                        # 9pt Arial, near enough for layout
CHART_H = 12.0                       # short enough to fit a laptop window
PLOT_Y, PLOT_H = 0.04, 0.76          # shared plot band, so the two charts align
ROW_PT = 15.0                        # Excel's default row height, in points
ROW_CM = ROW_PT / 72 * 2.54          # 0.529 cm -- not 0.4
LEGEND_ROW_PT = 30.0                 # uniform, so every swatch is the same size
CHART_TOP_ROW = 4                    # title on row 1, subtitle on row 2


def _cols_to_cm(*widths) -> float:
    """Excel column width (characters) -> centimetres."""
    px = sum(round(w * 7) + 5 for w in widths)
    return round(px / 96 * 2.54, 3)


AXIS_W = _cols_to_cm(AXIS_COL_W)


def _span_days(index) -> float:
    """Elapsed time covered by the plot, in days."""
    if len(index) < 2:
        return 0.0
    return (index[-1] - index[0]).total_seconds() / 86400.0


def _plot_width_detail(span_days: float, cm_per_day: float = CM_PER_DAY):
    """(drawn width, requested width, clamp) -- clamp is None, 'min' or 'max'."""
    want = span_days * cm_per_day
    if want > MAX_W:
        return MAX_W, want, "max"
    if want < MIN_W:
        return MIN_W, want, "min"
    return round(want, 1), want, None


def _plot_width(span_days: float, cm_per_day: float = CM_PER_DAY) -> float:
    return _plot_width_detail(span_days, cm_per_day)[0]


def _drawn_cm_per_day(span_days: float, cm_per_day: float = CM_PER_DAY) -> float:
    width = _plot_width(span_days, cm_per_day)
    return width / span_days if span_days else float("inf")


def chart_scale_note(result, cm_per_day: float = CM_PER_DAY) -> str | None:
    """A sentence when the chart is NOT drawn at the requested scale, else None.

    MIN_W and MAX_W are guards -- one stops a short record being a sliver, the
    other stops Excel choking on an absurd object -- but when either binds, the
    chart is no longer at the scale that was asked for. Staying quiet about
    that would leave two workbooks both claiming the same cm/day and not
    actually comparable, which is the one thing the control exists to provide.
    """
    span = _span_days(result.data.index)
    width, want, clamp = _plot_width_detail(span, cm_per_day)
    if clamp is None:
        return None
    drawn = width / span if span else float("inf")
    if clamp == "max":
        return (f"Chart scale reduced to fit: {cm_per_day:g} cm/day across "
                f"{span:.2f} days needs {want:,.0f} cm, beyond the "
                f"{MAX_W:,.0f} cm limit. Drawn at {drawn:.2f} cm/day, so "
                f"points sit closer together than the setting implies.")
    return (f"Chart scale raised to fit: {cm_per_day:g} cm/day across "
            f"{span:.2f} days is only {want:.1f} cm, under the {MIN_W:.0f} cm "
            f"minimum. Drawn at {drawn:.2f} cm/day, so the chart is more "
            f"magnified than the setting implies.")


def _layout(x, w):
    return Layout(manualLayout=ManualLayout(
        xMode="edge", yMode="edge", x=x, y=PLOT_Y, w=w, h=PLOT_H))


def _chrome(ch):
    """Square corners, no chart-area border, no fill -- sits flat on the sheet."""
    ch.roundedCorners = False
    ch.graphical_properties = GraphicalProperties(
        noFill=True, ln=LineProperties(noFill=True))
    ch.plot_area.graphicalProperties = GraphicalProperties(
        noFill=True, ln=LineProperties(noFill=True))
    return ch


def _axis_chart(source_ws, col_ix, nrows, y_title, frame, y_cols, x_lo,
                hexes):
    """
    The y-axis only. Sits in the frozen columns so it stays put while the wide
    plot scrolls. The series are parked off to the side of the data so nothing
    draws; they exist because a chart with no series renders no axis.
    """
    ch = ScatterChart()
    ch.title = None
    ch.legend = None
    ch.height, ch.width = CHART_H, AXIS_W
    _add_line_series(ch, source_ws, col_ix, nrows, hexes)
    _style_common(ch)
    _style_value_axis(ch, frame, y_cols, y_title, pin_left=True)
    ch.x_axis.delete = True
    ch.x_axis.scaling.min = round(x_lo - 10, 6)
    ch.x_axis.scaling.max = round(x_lo - 9.99, 6)
    ch.layout = _layout(0.84, 0.15)
    return _chrome(ch)


def _plot_chart(source_ws, col_ix, nrows, index, frame, y_cols, width,
                hexes):
    """The wide, scrolling half. No title, no legend, no y-axis labels."""
    ch = ScatterChart()
    ch.title = None
    ch.height, ch.width = CHART_H, width
    _add_line_series(ch, source_ws, col_ix, nrows, hexes)
    _style_common(ch)
    ch.legend = None
    _style_value_axis(ch, frame, y_cols, None)
    _style_time_axis(ch, index, target_ticks=max(6, int(width / 7)))
    # openpyxl's NoneSet maps the string "none" to None and then omits the
    # attribute, so tickLblPos can't suppress these -- hide the whole axis
    # instead. The valAx element is still written, so its gridlines survive.
    ch.y_axis.delete = True
    ch.y_axis.title = None
    ch.layout = _layout(0.004, 0.992)
    return _chrome(ch)


def _write_header(ws, title, subtitle, col):
    """Title and subtitle sit beside the frozen axis, so they scroll with it."""
    letter = get_column_letter(col)
    ws[f"{letter}1"] = title
    ws[f"{letter}1"].font = Font(name="Arial", size=12, bold=True)
    ws[f"{letter}2"] = subtitle
    ws[f"{letter}2"].font = Font(name="Arial", size=9, color="808080")
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 13


def _cell_legend(ws, cols, units, mixed, row, start_col, geometry=None,
                 colors=None):
    """
    A conventional horizontal legend, laid out in cells below the plot and to
    the right of the frozen axis, so it scrolls with the chart. Each entry gets
    a narrow swatch column plus however many default-width columns its label
    needs, so nothing wraps and nothing collides.
    """
    colors = colors or identity.series_colors(cols)
    ws.row_dimensions[row].height = LEGEND_ROW_PT
    col = start_col
    for i, c in enumerate(cols):
        label = identity.legend_label(c, units.get(c, ""), mixed,
                                      (geometry or {}).get(c, ""))
        ws.column_dimensions[get_column_letter(col)].width = SWATCH_W
        sw = ws.cell(row=row, column=col)
        # The swatch must match the line it stands for, on every sheet.
        sw.fill = PatternFill("solid",
                              fgColor=colors[c] if c in colors
                              else identity.SERIES_COLORS[
                                  i % len(identity.SERIES_COLORS)])

        cell = ws.cell(row=row, column=col + 1, value=label)
        cell.font = Font(name="Arial", size=9)
        cell.alignment = Alignment(vertical="center", indent=1)

        span = max(1, int(np.ceil((len(label) * CHAR_PX + 14) / DEFAULT_COL_PX)))
        col += 1 + span + 1          # swatch + label + a gap


# A panel is a title row, the chart, a legend row, then a gap before the next.
CHART_ROWS = int(np.ceil(CHART_H / ROW_CM))
PANEL_GAP_ROWS = 3

# What to call a unit on an axis when the bare symbol would not read as one.
UNIT_AXIS_LABELS = {
    "": "unit not detected",
    "1": "dimensionless",
    "qc_flag": "QARTOD flag",
}


def _unit_groups(result, cols):
    """[(axis label, [column, ...]), ...] -- one entry per data type.

    Series share a y-axis when, and only when, they share a unit. Two scales
    on one frame make any two series look however you want, so the grouping
    key is the CANONICAL unit: `m s-1` and `m/s` are one group, because they
    are one unit spelled two ways by two feeds.

    Derived columns are kept apart even when the unit matches. The
    stratification index is degC, but it is a temperature DIFFERENCE, not a
    temperature, and putting it on the temperature axis is the same mistake as
    a second axis -- one frame, two quantities.

    Order follows first appearance in `cols`, so the panels come out in the
    order the series were selected.
    """
    derived = set(getattr(result, "derived", []) or [])
    groups: dict[tuple, list] = {}
    for c in cols:
        info = (getattr(result, "columns", {}) or {}).get(c)
        unit = sk.canonical_unit(result.units.get(c, ""),
                                 getattr(info, "column", None) or c)
        groups.setdefault((c in derived, unit), []).append(c)
    out = []
    for (is_derived, unit), gcols in groups.items():
        label = UNIT_AXIS_LABELS.get(unit, unit)
        out.append((f"{label} (derived)" if is_derived else label, gcols))
    return out


def _scatter_pair(result, cols):
    """(x, y), unit -- the first two series sharing a unit, or (None, None).

    A scatter asserts that two quantities can be plotted against one another.
    That is only true when they measure the same thing, so the pair is drawn
    from a single unit group rather than from the head of the selection.
    Derived columns are excluded for the same reason they get their own panel:
    a temperature difference is not a temperature.
    """
    for label, gcols in _unit_groups(result, cols):
        if len(gcols) >= 2 and not label.endswith("(derived)"):
            return (gcols[0], gcols[1]), label
    return None, None


def _panel_title(ws, row, label, gcols, col):
    letter = get_column_letter(col)
    ws[f"{letter}{row}"] = (f"{label}  \u00b7  {len(gcols)} "
                            f"series" if len(gcols) != 1 else f"{label}  \u00b7  1 series")
    ws[f"{letter}{row}"].font = Font(name="Arial", size=10, bold=True)
    ws.row_dimensions[row].height = 15


def _write_chart_sheets(wb, result, cols, data_ws, norm_ws,
                        cm_per_day: float = CM_PER_DAY):
    """chart_raw stacks one panel per data type; chart_zscore stays combined."""
    n = len(result.data)
    width = _plot_width(_span_days(result.data.index), cm_per_day)
    units = sorted({result.units.get(c, "") for c in cols if result.units.get(c)})
    mixed = len(units) > 1
    x_lo = _serial(result.data.index[0].tz_convert(LOCAL_TZ).replace(tzinfo=None))
    lo = result.data.index[0].tz_convert(LOCAL_TZ)
    hi = result.data.index[-1].tz_convert(LOCAL_TZ)
    legend_row = CHART_TOP_ROW + CHART_ROWS + 1
    geometry = getattr(result, "geometry", None)
    made = []

    # Sheet column of each series: the data sheet has two time columns before
    # the data, the z-score sheet only one.
    raw_ix = {c: 3 + i for i, c in enumerate(cols)}
    z_ix = {c: 2 + i for i, c in enumerate(cols)}

    groups = _unit_groups(result, cols)
    # One colour per series for the whole workbook, so a line keeps its
    # identity from the raw panel to the z-score sheet to the legends.
    colors = identity.series_colors(cols)

    subtitle = (f"{result.interval} {result.aggregation} \u00b7 "
                f"{lo:%Y-%m-%d %H:%M} to {hi:%Y-%m-%d %H:%M} local \u00b7 "
                f"{n:,} intervals \u00b7 {len(cols)} series \u00b7 "
                f"{len(groups)} data type{'s' if len(groups) != 1 else ''}")

    # ---- chart_raw: one stacked panel per data type ------------------------
    # Each panel is scaled to its own group. A single shared axis had to span
    # every unit at once, which is what made it unreadable as soon as a second
    # data type was selected.
    ws = wb.create_sheet("chart_raw")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = AXIS_COL_W
    ws.freeze_panes = "B1"           # column A only -- the axis and nothing else
    _write_header(ws, "Raw values", subtitle, 2)

    row = CHART_TOP_ROW
    for label, gcols in groups:
        idx = [raw_ix[c] for c in gcols]
        _panel_title(ws, row, label, gcols, 2)
        top = row + 1
        ghex = [colors[c] for c in gcols]
        ws.add_chart(_axis_chart(data_ws, idx, n, label,
                                 result.data, gcols, x_lo, ghex), f"A{top}")
        ws.add_chart(_plot_chart(data_ws, idx, n, result.data.index,
                                 result.data, gcols, width, ghex), f"B{top}")
        # `mixed` is False within a panel: every series here shares a unit and
        # the panel title already states it, so repeating it on each entry is
        # noise.
        lrow = top + CHART_ROWS + 1
        _cell_legend(ws, gcols, result.units, False, lrow, 2, geometry, colors)
        row = lrow + PANEL_GAP_ROWS
    made.append("chart_raw")

    # ---- chart_zscore: deliberately NOT split ------------------------------
    # A z-score is unitless by construction, so every series already shares a
    # scale. Splitting it would defeat the only sheet that can compare shape
    # and timing across different units.
    zframe = (result.data - result.data.mean()) / result.data.std(ddof=0)
    ws = wb.create_sheet("chart_zscore")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = AXIS_COL_W
    ws.freeze_panes = "B1"
    _write_header(ws, "Standardised (z-score)", subtitle, 2)
    zidx = [z_ix[c] for c in cols]
    zhex = [colors[c] for c in cols]
    ws.add_chart(_axis_chart(norm_ws, zidx, n, "standard deviations",
                             zframe, cols, x_lo, zhex), f"A{CHART_TOP_ROW}")
    ws.add_chart(_plot_chart(norm_ws, zidx, n, result.data.index,
                             zframe, cols, width, zhex), f"B{CHART_TOP_ROW}")
    _cell_legend(ws, cols, result.units, mixed, legend_row, 2, geometry, colors)
    made.append("chart_zscore")

    # ---- stratification: its own panel, because it is a different quantity --
    made += _write_stratification_sheet(wb, result, cols, data_ws, subtitle,
                                        legend_row, cm_per_day)

    # ---- scatter: square, self-contained ----------------------------------
    # Only two series that SHARE A UNIT go on it. It used to take the first two
    # selected columns whatever they were, so picking temperature and wave
    # height produced a scatter of degC against metres -- a picture that will
    # happily show a trend line between two quantities that cannot have one.
    pair, punit = _scatter_pair(result, cols)
    if pair is not None:
        ws = wb.create_sheet("chart_scatter")
        ws.sheet_view.showGridLines = False
        xcol, ycol = pair
        xunit = result.units.get(xcol, "")
        yunit = result.units.get(ycol, "")
        _write_header(ws, f"{xcol} vs {ycol}", subtitle, 1)
        sc = ScatterChart()
        sc.title = None
        sc.height, sc.width = 15, 15
        # Reference the columns these two series actually occupy. They used to
        # be hardcoded to data columns 3 and 4, i.e. whichever two series were
        # selected first.
        xi, yi = 3 + cols.index(xcol), 3 + cols.index(ycol)
        ser = Series(Reference(data_ws, min_col=yi, min_row=2, max_row=n + 1),
                     Reference(data_ws, min_col=xi, min_row=2, max_row=n + 1))
        ser.smooth = False
        ser.marker = Marker(symbol="circle", size=4)
        ser.marker.graphicalProperties = GraphicalProperties(
            solidFill=colors[ycol], ln=LineProperties(noFill=True))
        ser.graphicalProperties = GraphicalProperties(ln=LineProperties(noFill=True))
        sc.series.append(ser)
        sc.legend = None
        _style_common(sc)
        _style_value_axis(sc, result.data, [ycol],
                          f"{ycol} [{yunit}]" if yunit else ycol)
        xv = result.data[xcol].to_numpy(dtype="float64")
        xv = xv[~np.isnan(xv)]
        if xv.size:
            pad = (xv.max() - xv.min()) * 0.08 or 1
            sc.x_axis.scaling.min = round(float(xv.min()) - pad, 2)
            sc.x_axis.scaling.max = round(float(xv.max()) + pad, 2)
        sc.x_axis.number_format = "0.0"
        _set_axis_title(sc.x_axis, f"{xcol} [{xunit}]" if xunit else xcol)
        ws.add_chart(_chrome(sc), f"A{CHART_TOP_ROW}")
        r = CHART_TOP_ROW + int(np.ceil(15 / ROW_CM)) + 1
        ws.cell(row=r, column=1, value=(
            f"Both axes are {UNIT_AXIS_LABELS.get(punit, punit)}. Series in "
            f"different units are not plotted against each other here: a "
            f"straight line through temperature and wave period would still "
            f"look like a relationship, and would not be one.")).font = NOTE
        made.append("chart_scatter")

    return made


def _write_stratification_sheet(wb, result, cols, data_ws, subtitle,
                                legend_row, cm_per_day: float = CM_PER_DAY):
    """A separate panel for SST(46254) - T(autoss).

    It gets its own sheet rather than a second axis on the raw chart: it is a
    temperature DIFFERENCE, not a temperature, and the no-dual-axis rule exists
    precisely so two different quantities never share one frame. Read it as the
    covariate -- when it is large the column is stratified, and a seabed logger
    and a near-surface sensor have no reason to agree.
    """
    derived = [c for c in getattr(result, "derived", []) if c in cols]
    if not derived:
        return []

    ws = wb.create_sheet("chart_stratification")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = AXIS_COL_W
    ws.freeze_panes = "B1"
    _write_header(ws, "Stratification index", subtitle, 2)

    n = len(result.data)
    # Same window, same scale -- this panel has to line up with chart_raw's.
    width = _plot_width(_span_days(result.data.index), cm_per_day)
    x_lo = _serial(result.data.index[0].tz_convert(LOCAL_TZ).replace(tzinfo=None))
    frame = result.data[derived]
    idx = [3 + cols.index(c) for c in derived]
    # Same colours as everywhere else -- the index appears on chart_raw too.
    colors = identity.series_colors(cols)
    dhex = [colors[c] for c in derived]

    ws.add_chart(_axis_chart(data_ws, idx, n,
                             "temperature difference [degC]", frame, derived,
                             x_lo, dhex), f"A{CHART_TOP_ROW}")
    ws.add_chart(_plot_chart(data_ws, idx, n, result.data.index,
                             frame, derived, width, dhex), f"B{CHART_TOP_ROW}")
    _cell_legend(ws, derived, result.units, False, legend_row, 2,
                 getattr(result, "geometry", None), colors)

    r = legend_row + 2
    ws.cell(row=r, column=2, value=(
        "Surface minus 5 m. Positive means the surface is warmer, i.e. a "
        "stratified water column. This is the covariate that says WHEN the "
        "seabed logger and the pier sensors should be expected to track each "
        "other -- not a sensor reading in its own right.")).font = NOTE
    return ["chart_stratification"]


def _write_provenance_sheet(wb, result, root, lag_reference=None, study=None,
                            cm_per_day: float = CM_PER_DAY):
    ws = wb.create_sheet("provenance")
    ws.cell(row=1, column=1, value="How this file was made").font = TITLE

    man = (study.manifest if study is not None else {}) or {}
    attached = man.get("source_files") or []
    val = man.get("validation") or {}

    rows = [
        ("generated (local)", datetime.now(LOCAL_TZ).strftime("%Y-%m-%d %H:%M:%S %Z")),
    ]
    if study is not None:
        rows += [
            ("study id", study.study_id),
            ("study label", study.label),
            ("study created (UTC)", study.created_utc),
            ("study status", study.status),
            ("study folder", str(Path(study.path).resolve())),
            ("pull window (UTC)",
             f"{(man.get('window_utc') or {}).get('start', '-')} to "
             f"{(man.get('window_utc') or {}).get('end', '-')}"
             f"  ({(man.get('window_utc') or {}).get('window_days', '-')} days)"),
            ("ingest sources",
             ", ".join((man.get("ingest") or {}).get("sources", [])) or "-"),
            ("tool version", man.get("tool_version", "-")),
            ("QC policy", (val.get("qc_policy")
                           or man.get("qc_policy") or "-")),
        ]
        rows += [("attached files",
                  f"{len(attached)} file(s)" if attached else "none")]
        for rec in attached:
            sha = (rec.get("sha256") or "-")[:16]
            rows.append((f"  {rec.get('file', '?')}",
                         f"station={rec.get('station', '?')}  "
                         f"rows={rec.get('n_rows', 0):,}  sha256={sha}..."))
            rows.append((f"    from", rec.get("original_path", "-")))
            for sh in rec.get("sheets", []):
                # State the time basis explicitly. For an attached file the
                # header is the only evidence there is, and for an ambiguous
                # one an assumption was made -- say which.
                basis = sh.get("time_kind", "?")
                if sh.get("assumed_timezone"):
                    basis += f" -> {sh['assumed_timezone']}"
                if sh.get("time_kind") == "ambiguous":
                    basis += "  (ASSUMED: the column named no zone)"
                rows.append((f"    sheet {sh.get('sheet', '?')}",
                             f"time={sh.get('time_column', '?')}  {basis}"))
    else:
        rows += [("source folder", str((root / "sources").resolve()))]

    span = _span_days(result.data.index)
    rows += [
        ("interval", result.interval),
        ("aggregation", result.aggregation),
        ("overlap rule", result.overlap),
        ("min samples per bin", result.min_samples),
        ("rows written", len(result.data)),
        # Presentation, but it belongs here: two workbooks of the same series
        # are only comparable by eye if they were drawn at the same scale --
        # so the scale actually DRAWN is recorded, not just the one requested.
        # 2 dp, not 3: the drawn figure is derived from a width rounded to
        # 0.1 cm for Excel, and at 3 dp that rounding shows up as a spurious
        # 3.6 -> 3.601 that reads as a scale change when nothing changed.
        ("chart scale requested (cm/day)", round(cm_per_day, 2)),
        ("chart scale drawn (cm/day)", round(_drawn_cm_per_day(span, cm_per_day), 2)),
        ("chart plot width (cm)", _plot_width(span, cm_per_day)),
        ("window start (local)",
         result.data.index[0].tz_convert(LOCAL_TZ).strftime("%Y-%m-%d %H:%M")
         if len(result.data) else "-"),
        ("window end (local)",
         result.data.index[-1].tz_convert(LOCAL_TZ).strftime("%Y-%m-%d %H:%M")
         if len(result.data) else "-"),
        ("lag reference", lag_reference or "-"),
        ("timezone", "America/Los_Angeles via zoneinfo; stored UTC internally"),
        ("series dropped",
         "; ".join(result.dropped) if result.dropped else "none"),
    ]
    scale_note = chart_scale_note(result, cm_per_day)
    if scale_note:
        rows.append(("chart scale WARNING", scale_note))
    r = 3
    for k, v in rows:
        ws.cell(row=r, column=1, value=k).font = HEAD
        cell = ws.cell(row=r, column=2, value=v)
        cell.font = BODY
        if k.endswith("WARNING"):
            cell.fill = WARNFILL
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    # ---- clock check verdicts ---------------------------------------------
    checks = val.get("clock_checks") or []
    if checks:
        r += 1
        ws.cell(row=r, column=1, value="Clock verification").font = TITLE
        r += 1
        for j, h in enumerate(["signal", "role", "observed peak (UTC h)",
                               "expected (UTC h)", "offset (h)", "amplitude",
                               "verdict"], start=1):
            ws.cell(row=r, column=j, value=h)
        _style_header(ws, row=r, ncols=7)
        ws.freeze_panes = None
        for c in checks:
            r += 1
            ws.cell(row=r, column=1, value=c.get("signal", "-")).font = BODY
            ws.cell(row=r, column=2, value=c.get("role", "-")).font = BODY
            ws.cell(row=r, column=3, value=c.get("observed_peak_hour_utc")).font = BODY
            ws.cell(row=r, column=4, value=c.get("expected_peak_hour_utc")).font = BODY
            ws.cell(row=r, column=5, value=c.get("offset_hours")).font = BODY
            ws.cell(row=r, column=6, value=c.get("amplitude")).font = BODY
            cell = ws.cell(row=r, column=7, value="OK" if c.get("ok") else "FAIL")
            cell.font = BODY
            if not c.get("ok"):
                cell.fill = WARNFILL
        r += 2
        ws.cell(row=r, column=1, value=(
            "A timestamp column is only UTC if a signal with a known solar phase "
            "says so. Air temperature peaks ~2 h after local solar noon; the S2 "
            "pressure tide peaks ~10:00 and ~22:00 local solar. See "
            "ingest/clockcheck.py.")).font = NOTE
        r += 1

    # ---- per-series detail -------------------------------------------------
    r += 1
    ws.cell(row=r, column=1, value="Series").font = TITLE
    r += 1
    heads = ["output column", "unit", "depth / frame", "time basis",
             "QARTOD-3 kept", "source", "native cadence"]
    for j, h in enumerate(heads, start=1):
        ws.cell(row=r, column=j, value=h)
    _style_header(ws, row=r, ncols=len(heads))
    ws.freeze_panes = None

    suspect = {}
    for s in (man.get("series") or []):
        suspect[f"{s.get('station')}.{s.get('variable')}"] = s.get("n_flagged_suspect")

    for c in result.data.columns:
        r += 1
        ws.cell(row=r, column=1, value=c).font = BODY
        ws.cell(row=r, column=2, value=result.units.get(c, "")).font = BODY
        ws.cell(row=r, column=3,
                value=getattr(result, "geometry", {}).get(c, "")).font = BODY
        basis = getattr(result, "time_trust", {}).get(c, "")
        cell = ws.cell(row=r, column=4, value=basis)
        cell.font = BODY
        if "unverified" in basis.lower():
            cell.fill = WARNFILL
        ws.cell(row=r, column=5,
                value=suspect.get(c.split(" ")[0], "")).font = BODY
        ws.cell(row=r, column=6, value=result.sources.get(c, "")).font = BODY
        ws.cell(row=r, column=7, value=result.cadences.get(c, "")).font = BODY

    r += 2
    for line in [
        "Assumptions applied on load:",
        "  - A column named 'time_utc' is UTC, verified by ingest/clockcheck.py",
        "    against the solar phase of air temperature and barometric pressure.",
        "  - A legacy column headed 'time (UTC)' is loaded but marked UNVERIFIED",
        "    and shaded above. In this study's original workbook that column",
        "    held Pacific local time, not UTC -- a column name is not evidence of",
        "    a zone. No offset is applied to it here; guessing one is what caused",
        "    the original error.",
        "  - Columns headed 'Date-Time (PDT)' are logger-local wall time and are",
        "    converted via zoneinfo, never by adding a constant.",
        "  - Fahrenheit columns are converted to Celsius; the original unit is",
        "    recorded above.",
        "  - Rows whose timestamp is ambiguous or nonexistent across a DST",
        "    transition are dropped rather than guessed.",
        "  - QARTOD flags 4 (fail) and 9 (missing) are rejected; 1, 2 and 3 pass,",
        "    and the count of suspect (3) values KEPT is listed per series above.",
        "  - BinKey* columns in old snapshots are ignored; binning is done here,",
        "    at the interval named above.",
    ]:
        ws.cell(row=r, column=1, value=line).font = NOTE
        r += 1

    _autosize(ws, maxw=60)
    return ws


def write_workbook(result, root: Path, out_path: Path,
                   lag_table=None, lag_reference=None, study=None,
                   cm_per_day: float | None = None) -> Path:
    cm_per_day = CM_PER_DAY if not cm_per_day else float(cm_per_day)
    wb = Workbook()
    wb.remove(wb.active)

    data_ws, cols = _write_data_sheet(wb, result)
    norm_ws = _write_zscore_sheet(wb, result, cols)
    _write_counts_sheet(wb, result, cols)
    _write_stats_sheet(wb, result, cols, lag_table, lag_reference)
    charts = _write_chart_sheets(wb, result, cols, data_ws, norm_ws, cm_per_day)
    _write_provenance_sheet(wb, result, root, lag_reference, study, cm_per_day)

    order = ([wb[c] for c in charts]
             + [wb["data"], wb["stats"], wb["counts"],
                wb["normalized"], wb["provenance"]])
    wb._sheets = order
    wb.active = 0

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path = unique_path(out_path)
    wb.save(out_path)
    return out_path


def unique_path(path: Path) -> Path:
    """`path`, or the first `name-2`, `name-3`... that does not exist yet.

    A generated workbook must never replace one that is already there. The
    output name carries only the first three series, the interval and a stamp
    good to the MINUTE -- it says nothing about the aggregation, the overlap
    rule, the minimum sample count, the window or the chart scale. So the whole
    point of generating twice in a minute is to vary one of the things the name
    does not record, and that is exactly when the second file would have landed
    on the first.

    This also sidesteps writing to a workbook Excel currently has open, which
    fails outright rather than quietly.
    """
    path = Path(path)
    if not path.exists():
        return path
    for i in range(2, 1000):
        candidate = path.with_name(f"{path.stem}-{i}{path.suffix}")
        if not candidate.exists():
            return candidate
    raise FileExistsError(f"cannot find a free name beside {path}")


def default_output_name(cols, interval) -> str:
    stamp = datetime.now(LOCAL_TZ).strftime("%Y%m%d_%H%M")
    def slug(c):
        return re.sub(r"[^A-Za-z0-9]+", "", c.split(".")[-1])[:12] or "series"
    short = "_".join(slug(c) for c in cols[:3])
    more = f"_plus{len(cols) - 3}" if len(cols) > 3 else ""
    return f"compare_{short}{more}_{interval}_{stamp}.xlsx"
