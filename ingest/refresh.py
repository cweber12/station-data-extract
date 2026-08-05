"""
refresh.py -- refresh the Power Query workbook through Excel COM and snapshot it.

OPTIONAL. The Python ingest in erddap.py / coops.py / ndbc_realtime.py is the
primary path and needs no Excel at all. This module exists so a project can also
carry a refreshed .xlsx for people who work in the workbook directly.

Things that will otherwise cost you an afternoon:

  * `DispatchEx`, never `Dispatch`. `Dispatch` attaches to the user's already-
    open Excel and will fight them for control of it.
  * `RefreshAll()` is ASYNCHRONOUS. Setting BackgroundQuery=False on every OLEDB
    connection first, then calling CalculateUntilAsyncQueriesDone(), is what
    makes it behave. Without that you SaveAs a workbook full of stale data and
    everything downstream looks fine while being wrong.
  * A `~$name.xlsx` lock file next to the source means the workbook is open (or
    was, and Excel crashed). Refreshing then either fails or silently saves a
    read-only copy. We refuse up front with a clear message. Orphaned locks have
    already been observed in this repo, so this is a real failure mode.
  * PRIVACY LEVELS. A first run may block on "Information is required about data
    privacy". This cannot be set reliably over COM. One-time manual fix:
        Data -> Get Data -> Query Options -> Privacy
        -> "Always ignore Privacy Level settings"
    With Visible=False such a prompt hangs forever, which is why there is a
    wall-clock timeout and why Excel is always quit in a finally block.
  * SaveAs goes to the PROJECT folder. Never back over sources/.
"""

from __future__ import annotations

import datetime as dt
import shutil
import threading
from pathlib import Path

XL_OPENXML_WORKBOOK = 51        # xlOpenXMLWorkbook
DEFAULT_TIMEOUT_S = 600


class ExcelNotAvailable(RuntimeError):
    """Excel or pywin32 is missing. The Python ingest does not need either."""


class WorkbookLocked(RuntimeError):
    """A ~$ lock file exists, or the file is open. Refuse rather than guess."""


def _lock_file(path: Path) -> Path:
    return path.parent / f"~${path.name}"


def check_not_locked(src: Path) -> None:
    src = Path(src)
    lock = _lock_file(src)
    if lock.exists():
        raise WorkbookLocked(
            f"{lock.name} exists beside {src.name}.\n"
            f"The workbook is open in Excel, or Excel crashed and left the lock "
            f"behind.\nClose the workbook (or delete {lock}) and try again."
        )
    try:
        with src.open("r+b"):
            pass
    except PermissionError as e:
        raise WorkbookLocked(
            f"{src.name} is locked by another process. Close it in Excel first."
        ) from e


def excel_available() -> bool:
    try:
        import win32com.client as win32
    except ImportError:
        return False
    try:
        xl = win32.DispatchEx("Excel.Application")
    except Exception:
        return False
    try:
        xl.Quit()
    except Exception:
        pass
    return True


def read_meta_refresh(xlsx: Path) -> dict:
    """Pull the meta_refresh key/value sheet out of a workbook, if present.

    Absent sheet -> empty dict, and the caller marks the project 'incomplete'
    rather than inventing a fetch time.
    """
    from openpyxl import load_workbook
    try:
        wb = load_workbook(xlsx, read_only=True, data_only=True)
    except Exception:
        return {}
    try:
        name = next((s for s in wb.sheetnames if s.lower() == "meta_refresh"), None)
        if name is None:
            return {}
        ws = wb[name]
        out = {}
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip()
            if key.lower() in ("key", ""):
                continue
            out[key] = None if len(row) < 2 or row[1] is None else str(row[1]).strip()
        return out
    finally:
        wb.close()


def refresh_and_snapshot(src_xlsx: Path, dest_xlsx: Path,
                         timeout_s: int = DEFAULT_TIMEOUT_S) -> dict:
    """Refresh all Power Query connections in src_xlsx and SaveAs to dest_xlsx.

    Never writes back to src_xlsx. Returns the meta_refresh key/value dict.
    """
    src, dest = Path(src_xlsx).resolve(), Path(dest_xlsx).resolve()
    if src == dest:
        raise ValueError("refusing to save over the source workbook")
    if "sources" in [p.name for p in dest.parents]:
        raise ValueError("refusing to write into sources/ -- it is read-only")

    try:
        import win32com.client as win32
    except ImportError as e:
        raise ExcelNotAvailable(
            "pywin32 is not installed. Install it, or create the project with "
            "refresh disabled -- the Python ingest does not need Excel."
        ) from e

    check_not_locked(src)
    dest.parent.mkdir(parents=True, exist_ok=True)

    result: dict = {}
    error: list[BaseException] = []

    def work():
        xl = wb = None
        # The work runs on a worker thread so the wall-clock timeout can win a
        # race against a modal privacy prompt. Every COM apartment needs its own
        # CoInitialize; without this the first Dispatch raises
        # "CoInitialize has not been called."
        import pythoncom
        pythoncom.CoInitialize()
        try:
            xl = win32.DispatchEx("Excel.Application")
            xl.Visible = False
            xl.DisplayAlerts = False
            xl.AskToUpdateLinks = False
            wb = xl.Workbooks.Open(str(src), UpdateLinks=0, ReadOnly=False)

            # Force synchronous refresh. Not every connection type exposes
            # OLEDBConnection, so this is best-effort per connection.
            for conn in wb.Connections:
                try:
                    conn.OLEDBConnection.BackgroundQuery = False
                except Exception:
                    pass

            wb.RefreshAll()
            xl.CalculateUntilAsyncQueriesDone()

            wb.SaveAs(str(dest), FileFormat=XL_OPENXML_WORKBOOK)
            wb.Close(SaveChanges=False)
            wb = None
        except BaseException as e:          # noqa: BLE001 - re-raised on the caller
            error.append(e)
        finally:
            try:
                if wb is not None:
                    wb.Close(SaveChanges=False)
            except Exception:
                pass
            try:
                if xl is not None:
                    xl.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    t = threading.Thread(target=work, daemon=True)
    t.start()
    t.join(timeout_s)
    if t.is_alive():
        raise TimeoutError(
            f"Excel did not finish within {timeout_s}s. The usual cause is a "
            f"privacy-level prompt, which is invisible with Visible=False. Open "
            f"{src.name} once by hand and set Data -> Get Data -> Query Options "
            f"-> Privacy -> 'Always ignore Privacy Level settings'."
        )
    if error:
        raise error[0]
    if not dest.is_file():
        raise RuntimeError(f"Excel reported success but {dest} was not written")

    result = read_meta_refresh(dest)
    return result


def copy_snapshot(src_xlsx: Path, dest_xlsx: Path) -> dict:
    """The no-Excel path: copy the workbook as-is, claim nothing about freshness."""
    src, dest = Path(src_xlsx), Path(dest_xlsx)
    dest.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dest)
    return read_meta_refresh(dest)


def _main(argv=None):
    import argparse
    ap = argparse.ArgumentParser(description="Refresh and snapshot the workbook.")
    ap.add_argument("src", type=Path)
    ap.add_argument("dest", type=Path)
    ap.add_argument("--timeout", type=int, default=DEFAULT_TIMEOUT_S)
    ap.add_argument("--check", action="store_true",
                    help="only report whether Excel is usable")
    args = ap.parse_args(argv)

    if args.check:
        print("Excel available:", excel_available())
        print("locked:", end=" ")
        try:
            check_not_locked(args.src)
            print("no")
        except WorkbookLocked as e:
            print(f"YES\n{e}")
        return 0

    meta = refresh_and_snapshot(args.src, args.dest, args.timeout)
    print(f"wrote {args.dest}")
    for k, v in meta.items():
        print(f"  {k:16} {v}")
    return 0


if __name__ == "__main__":
    raise SystemExit(_main())
