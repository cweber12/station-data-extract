"""
sensorkit.py -- data layer for the La Jolla sensor comparison tool.

No GUI code here on purpose: everything in this module is importable and
testable from a plain Python prompt. compare.py is a thin Tkinter shell on top.

Design rules baked in:
  * A project is read-only input. outputs/ is never scanned. There is no path
    by which a generated file becomes an input.
  * Every timestamp is converted to UTC on load -- but a column NAME is never
    accepted as evidence of a timezone. See the time contract below.
  * Every temperature is converted to degC on load. The original unit is
    recorded and carried through to the provenance sheet.
  * Resampling happens at query time. Nothing is pre-binned.

THE TIME CONTRACT
-----------------
This module used to put `time (utc)` first in TIME_PREFERENCE, with a comment
saying "UTC always wins so we never have to trust a local label". That comment
is what the bug hid behind, and it was exactly backwards for this workbook.

The ERDDAP-derived `time (UTC)` columns in ja_jolla_sensors.xlsx contain PACIFIC
LOCAL TIME. Power Query's implicit text -> datetime conversion applied the
machine's UTC-7 offset and discarded the zone. The proof is in the request: the
workbook asked ERDDAP for `time>=2026-06-18T00:00:00Z` and the stored column
begins at `2026-06-17 17:00`, seven hours before a bound the server enforced in
true UTC. A harmonic fit agrees -- air temperature at Scripps Pier peaked at
14.45 in that column, which is local solar afternoon, not 21.8 UTC.

So: a column name is not evidence of a zone. Only ingest/clockcheck.py is.
Data that came through the Python ingest carries `time_utc` and is trusted.
A legacy `time (UTC)` column with no `time_utc` beside it is loaded but marked
UNVERIFIED, surfaced in amber in the UI, and labelled as such on the provenance
sheet. It is not silently corrected -- guessing an offset is how the first bug
happened.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd
from openpyxl import load_workbook

LOCAL_TZ = ZoneInfo("America/Los_Angeles")

SOURCES_DIRNAME = "sources"
OUTPUTS_DIRNAME = "outputs"

# Trust levels for a table's time column.
TIME_VERIFIED = "verified"        # time_utc from the Python ingest
TIME_UNVERIFIED = "unverified"    # legacy `time (UTC)`, provenance unknown
TIME_LOCAL = "local"              # honestly-labelled local wall time

UNVERIFIED_NOTE = ("UTC (unverified -- legacy column, see AGENT_TASK.md 0.1 and "
                   "the time contract in sensorkit.py)")

# Time columns, in the order we prefer them.
#
# `time_utc` is first because it is the only one whose zone has been VERIFIED --
# it is produced by ingest/, which parses ISO-8601 with an explicit Z and asserts
# tz-awareness. Everything below it is legacy and carries a trust level.
TIME_PREFERENCE = [
    (re.compile(r"^time_utc$", re.I), "UTC", TIME_VERIFIED),
    (re.compile(r"time\s*\(utc\)", re.I), "UTC", TIME_UNVERIFIED),
    (re.compile(r"\butc\b", re.I), "UTC", TIME_UNVERIFIED),
    (re.compile(r"time\s*\(local\)", re.I), "LOCAL", TIME_UNVERIFIED),
    (re.compile(r"time\s*\(pdt\)", re.I), "LOCAL", TIME_UNVERIFIED),
    (re.compile(r"date-?\s*time", re.I), "LOCAL", TIME_LOCAL),
    (re.compile(r"^time$", re.I), "LOCAL", TIME_UNVERIFIED),
]

# Unit sniffing from column headers. First match wins.
UNIT_PATTERNS = [
    (re.compile(r"°\s*F|\bdeg(ree)?[_ ]?F\b", re.I), "degF"),
    (re.compile(r"°\s*C|\bdeg(ree)?[_ ]?C(elsius)?\b", re.I), "degC"),
    (re.compile(r"\(m\s*s-1\)", re.I), "m/s"),
    (re.compile(r"\(hPa\)", re.I), "hPa"),
    (re.compile(r"\(NTU\)", re.I), "NTU"),
    (re.compile(r"\(mg\.L-1\)", re.I), "mg/L"),
    (re.compile(r"\(microg\.L-1\)", re.I), "ug/L"),
    (re.compile(r"degrees_true", re.I), "deg"),
    (re.compile(r"\(1e-3\)", re.I), "PSU"),
    (re.compile(r"\(m\)", re.I), "m"),
    (re.compile(r"\(s\)", re.I), "s"),
    (re.compile(r"water\s*level", re.I), "ft"),
    (re.compile(r"_qc_agg$", re.I), "qc_flag"),
]

# The QC-screened `*_ok` columns carry no unit in their header, so map them
# explicitly. Add a line here when a new derived column appears.
UNIT_OVERRIDES = {
    "wtmp_ok": "degC", "sal_ok": "PSU", "ph_ok": "pH",
    "do_ok": "mg/L", "chl_ok": "ug/L", "turb_ok": "NTU",
}

# Columns that are bookkeeping, not data. BinKey* are the precomputed bin
# columns the audit flagged -- resampling replaces them entirely. The Power
# Query that produced them is gone, but OLD SNAPSHOTS STILL CONTAIN THEM, so
# this filter stays as a safety net. `station` is numeric for 46254 and text
# elsewhere, and is an identifier either way.
NOISE_COLUMNS = re.compile(
    r"^(Unnamed:|#$|BinKey\d+$|station$|wd_sin$|wd_cos$|mwd_sin$|mwd_cos$)", re.I)

AGGREGATIONS = {
    "mean": "mean",
    "median": "median",
    "min": "min",
    "max": "max",
    "std": "std",
    "count": "count",
}

INTERVALS = [
    "5min", "10min", "15min", "20min", "30min",
    "1h", "2h", "3h", "6h", "12h", "1D",
]


# --------------------------------------------------------------------------
# catalogue
# --------------------------------------------------------------------------

@dataclass
class ColumnInfo:
    file: str
    table: str
    column: str
    kind: str            # "time" | "data"
    unit: str            # detected unit, or "" if unknown
    n_nonnull: int
    n_rows: int
    # Geometry, joined from config/stations.yaml -- never from the feed, which
    # reports z = 0.0 for both the pier station and the surface buoy.
    depth_m: float | None = None
    reference_frame: str = "unknown"
    station: str | None = None
    variable: str | None = None

    @property
    def key(self) -> str:
        return f"{self.file}::{self.table}::{self.column}"

    @property
    def label(self) -> str:
        """Short name used as the output column header."""
        return f"{self.table}.{self.column}"

    @property
    def coverage(self) -> float:
        return 0.0 if not self.n_rows else self.n_nonnull / self.n_rows

    @property
    def depth_label(self) -> str:
        """'3.43 m' / 'bed' / '?' -- the depth half of a legend entry."""
        if self.depth_m is None:
            return "bed" if self.reference_frame == "earth" and \
                self.station == "yellow_buoy" else "?"
        return f"{self.depth_m:.2f}".rstrip("0").rstrip(".") + " m"

    @property
    def geometry_label(self) -> str:
        return f"({self.depth_label}, {self.reference_frame})"


@dataclass
class TableInfo:
    file: str
    name: str
    sheet: str
    n_rows: int
    time_column: str | None
    time_basis: str | None          # "UTC" or "LOCAL"
    columns: list[ColumnInfo] = field(default_factory=list)
    time_trust: str = TIME_UNVERIFIED
    path: Path | None = None        # absolute location, so loading needs no root
    kind: str = "xlsx"              # "xlsx" | "parquet"
    project_label: str | None = None   # set when comparing two projects

    @property
    def data_columns(self) -> list[ColumnInfo]:
        return [c for c in self.columns if c.kind == "data"]

    @property
    def time_is_verified(self) -> bool:
        return self.time_trust in (TIME_VERIFIED, TIME_LOCAL)

    @property
    def time_basis_note(self) -> str:
        if self.time_trust == TIME_VERIFIED:
            return "UTC (verified by ingest/clockcheck.py)"
        if self.time_trust == TIME_LOCAL:
            return "America/Los_Angeles wall time (honestly labelled)"
        if self.time_basis == "UTC":
            return UNVERIFIED_NOTE
        return "local wall time (unverified label)"


def detect_unit(header: str) -> str:
    key = header.strip()
    if key in UNIT_OVERRIDES:
        return UNIT_OVERRIDES[key]
    for pat, unit in UNIT_PATTERNS:
        if pat.search(header):
            return unit
    return ""


def detect_time_column(headers) -> tuple[str | None, str | None, str]:
    """Return (column, basis, trust). Preference order encodes the trust rule."""
    for pat, basis, trust in TIME_PREFERENCE:
        for h in headers:
            if isinstance(h, str) and pat.search(h):
                return h, basis, trust
    return None, None, TIME_UNVERIFIED


def _clean_frame(df: pd.DataFrame) -> pd.DataFrame:
    keep = [c for c in df.columns
            if isinstance(c, str) and not NOISE_COLUMNS.match(c.strip())]
    df = df[keep]
    return df.dropna(how="all")


def _read_excel_table(path: Path, sheet: str, ref: str | None) -> pd.DataFrame:
    """Read one sheet, or one table range within a sheet."""
    if ref is None:
        df = pd.read_excel(path, sheet_name=sheet)
    else:
        m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", ref)
        first_row = int(m.group(2))
        df = pd.read_excel(
            path, sheet_name=sheet,
            skiprows=first_row - 1,
            usecols=f"{m.group(1)}:{m.group(3)}",
        )
    return _clean_frame(df)


def _read_table(path: Path, sheet: str, ref: str | None) -> pd.DataFrame:
    """Dispatch on suffix so parquet and xlsx share one loading path."""
    if Path(path).suffix.lower() == ".parquet":
        return pd.read_parquet(path)
    return _read_excel_table(path, sheet, ref)


def scan_workbook(path: Path, geometry: dict | None = None) -> list[TableInfo]:
    """Catalogue every table (or bare sheet) in a workbook."""
    wb = load_workbook(path, read_only=True)
    tables: list[TableInfo] = []

    for ws in wb.worksheets:
        refs = {}
        try:
            for tname in ws.tables:
                obj = ws.tables[tname]
                refs[tname] = obj.ref if hasattr(obj, "ref") else obj
        except Exception:
            refs = {}

        targets = list(refs.items()) if refs else [(ws.title, None)]

        for tname, ref in targets:
            try:
                df = _read_excel_table(path, ws.title, ref)
            except Exception:
                continue
            if df.empty or len(df.columns) < 2:
                continue

            tcol, basis, trust = detect_time_column(df.columns)
            if tcol is None:
                continue  # nothing to align on; not comparable

            # A legacy `time (UTC)` beside a real `time_utc` is redundant, not
            # suspicious -- the preference order already picked the good one.
            info = TableInfo(
                file=path.name, name=tname, sheet=ws.title,
                n_rows=len(df), time_column=tcol, time_basis=basis,
                time_trust=trust, path=Path(path), kind="xlsx",
            )
            station = _station_from_table_name(tname)
            geo = (geometry or {}).get(station or "", {})
            for col in df.columns:
                if col == tcol:
                    continue
                s = df[col]
                if not pd.api.types.is_numeric_dtype(s):
                    continue
                info.columns.append(ColumnInfo(
                    file=path.name, table=tname, column=str(col),
                    kind="data", unit=detect_unit(str(col)),
                    n_nonnull=int(s.notna().sum()), n_rows=len(df),
                    depth_m=geo.get("depth_m"),
                    reference_frame=geo.get("reference_frame", "unknown"),
                    station=station,
                ))
            if info.columns:
                tables.append(info)

    wb.close()
    return tables


def _station_from_table_name(name: str) -> str | None:
    """Best-effort station id from a legacy sheet name like `src_LJAC1`."""
    m = re.match(r"(?:src_)?([A-Za-z0-9_]+)$", str(name).strip())
    if not m:
        return None
    token = m.group(1)
    known = {"ljac1": "LJAC1", "ljpc1": "LJPC1", "46254": "46254",
             "yellow_buoy": "yellow_buoy", "waterlevel": "LJAC1",
             "9410230_wl": "LJAC1"}
    low = token.lower()
    if low in known:
        return known[low]
    if low.startswith("sccoos"):
        return "autoss"
    return None


def scan_parquet(path: Path, project_label: str | None = None
                 ) -> list[TableInfo]:
    """Catalogue a canonical long frame: one TableInfo per (station, variable).

    The long frame is the Python ingest's output. Its time column is `time_utc`,
    tz-aware UTC, produced by a parser that was given an explicit Z -- so every
    table from here is TIME_VERIFIED.
    """
    df = pd.read_parquet(path)
    if df.empty or "station" not in df or "variable" not in df:
        return []

    tables: list[TableInfo] = []
    for (station, variable), g in df.groupby(["station", "variable"], sort=True):
        unit = str(g["unit"].iloc[0]) if "unit" in g else ""
        depth = g["depth_m"].iloc[0] if "depth_m" in g else None
        frame = str(g["reference_frame"].iloc[0]) if "reference_frame" in g \
            else "unknown"
        name = f"{station}.{variable}"
        info = TableInfo(
            file=path.name, name=name, sheet="observations",
            n_rows=int(len(g)), time_column="time_utc", time_basis="UTC",
            time_trust=TIME_VERIFIED, path=Path(path), kind="parquet",
            project_label=project_label,
        )
        # `table` is the station alone, not `station.variable`, so that
        # ColumnInfo.label reads `LJAC1.sea_water_temperature` rather than
        # repeating the variable twice.
        info.columns.append(ColumnInfo(
            file=path.name, table=str(station), column=variable, kind="data",
            unit=unit, n_nonnull=int(g["value"].notna().sum()),
            n_rows=int(len(g)),
            depth_m=None if pd.isna(depth) else float(depth),
            reference_frame=frame, station=str(station), variable=str(variable),
        ))
        tables.append(info)
    return tables


def _geometry_map(root: Path | None) -> dict:
    """{station -> {depth_m, reference_frame}} from config/stations.yaml."""
    if root is None:
        return {}
    try:
        from ingest.config import load_config
        cfg = load_config(root)
    except Exception:
        return {}
    return {s.key: {"depth_m": s.depth_m, "reference_frame": s.reference_frame}
            for s in cfg.stations.values()}


def build_catalog(root: Path, *, sources_dirname: str = SOURCES_DIRNAME,
                  config_root: Path | None = None
                  ) -> dict[str, list[TableInfo]]:
    """Scan one directory of workbooks. outputs/ is deliberately never touched."""
    src = Path(root) / sources_dirname
    if not src.is_dir():
        raise FileNotFoundError(f"no {sources_dirname}/ directory under {root}")

    geometry = _geometry_map(config_root or root)
    catalog: dict[str, list[TableInfo]] = {}
    for p in sorted(src.glob("*.xlsx")):
        if p.name.startswith("~$"):
            continue
        tabs = scan_workbook(p, geometry)
        if tabs:
            catalog[p.name] = tabs
    return catalog


def build_catalog_project(project, config_root: Path | None = None,
                          label_prefix: str | None = None
                          ) -> dict[str, list[TableInfo]]:
    """Catalogue a project: cache/*.parquet first, then any workbook snapshot.

    `label_prefix` is set when two projects are open at once, so a series from
    each can carry its project label and stay distinguishable.
    """
    geometry = _geometry_map(config_root)
    catalog: dict[str, list[TableInfo]] = {}

    for p in sorted(project.cache_dir.glob("*.parquet")):
        tabs = scan_parquet(p, project_label=label_prefix)
        if tabs:
            key = f"{label_prefix}: {p.name}" if label_prefix else p.name
            catalog[key] = tabs

    if project.workbook_dir.is_dir():
        for p in sorted(project.workbook_dir.glob("*.xlsx")):
            if p.name.startswith("~$"):
                continue
            try:
                tabs = scan_workbook(p, geometry)
            except Exception:
                continue
            if tabs:
                for t in tabs:
                    t.project_label = label_prefix
                key = f"{label_prefix}: {p.name}" if label_prefix else p.name
                catalog[key] = tabs

    return catalog


# --------------------------------------------------------------------------
# loading and normalisation
# --------------------------------------------------------------------------

def to_utc(series: pd.Series, basis: str) -> pd.DatetimeIndex:
    """Attach a zone. `basis` decides; the column name never does."""
    idx = pd.to_datetime(series, errors="coerce")
    di = pd.DatetimeIndex(idx)
    if di.tz is not None:
        return di.tz_convert("UTC")
    if basis == "UTC":
        return di.tz_localize("UTC")
    # Wall-clock local. ambiguous/nonexistent land on NaT and get dropped,
    # which is the honest answer for the two hours a year that are undefined.
    return di.tz_localize(LOCAL_TZ, ambiguous="NaT",
                          nonexistent="NaT").tz_convert("UTC")


def convert_units(values: pd.Series, unit: str) -> tuple[pd.Series, str]:
    if unit == "degF":
        return (values - 32.0) * 5.0 / 9.0, "degC"
    return values, unit


_frame_cache: dict[tuple, pd.DataFrame] = {}


def load_series(table: TableInfo, column: ColumnInfo,
                convert: bool = True) -> tuple[pd.Series, str]:
    """Return one column as a UTC-indexed Series, plus its final unit."""
    path = table.path
    if path is None:
        raise ValueError(f"{table.name}: no path recorded on the table")

    if table.kind == "parquet":
        ck = ("parquet", str(path))
        if ck not in _frame_cache:
            _frame_cache[ck] = pd.read_parquet(path)
        full = _frame_cache[ck]
        g = full[(full["station"].astype(str) == str(column.station))
                 & (full["variable"].astype(str) == str(column.variable))]
        idx = to_utc(g["time_utc"], "UTC")
        s = pd.Series(pd.to_numeric(g["value"], errors="coerce").values,
                      index=idx, name=column.label)
    else:
        ck = ("xlsx", str(path), table.sheet, table.name)
        if ck not in _frame_cache:
            wb = load_workbook(path, read_only=True)
            ws = wb[table.sheet]
            ref = None
            try:
                if table.name in ws.tables:
                    obj = ws.tables[table.name]
                    ref = obj.ref if hasattr(obj, "ref") else obj
            except Exception:
                ref = None
            wb.close()
            _frame_cache[ck] = _read_excel_table(path, table.sheet, ref)

        df = _frame_cache[ck]
        idx = to_utc(df[table.time_column], table.time_basis)
        s = pd.Series(pd.to_numeric(df[column.column], errors="coerce").values,
                      index=idx, name=column.label)

    s = s[s.index.notna()].sort_index()
    s = s[~s.index.duplicated(keep="first")]

    unit = column.unit
    if convert:
        s, unit = convert_units(s, unit)
    return s.dropna(), unit


def native_cadence(s: pd.Series) -> pd.Timedelta | None:
    if len(s) < 3:
        return None
    d = pd.Series(s.index).diff().dropna()
    return d.mode().iloc[0] if len(d) else None


# --------------------------------------------------------------------------
# alignment
# --------------------------------------------------------------------------

@dataclass
class BuildResult:
    data: pd.DataFrame          # UTC-indexed, one column per series
    counts: pd.DataFrame        # samples that went into each cell
    units: dict[str, str]
    cadences: dict[str, str]
    sources: dict[str, str]
    interval: str
    aggregation: str
    overlap: str
    min_samples: int
    dropped: list[str] = field(default_factory=list)
    geometry: dict[str, str] = field(default_factory=dict)
    time_trust: dict[str, str] = field(default_factory=dict)
    columns: dict[str, ColumnInfo] = field(default_factory=dict)
    derived: list[str] = field(default_factory=list)


def series_label(table: TableInfo, col: ColumnInfo) -> str:
    """`baseline: LJAC1.sea_water_temperature` when two projects are open."""
    base = col.label
    return f"{table.project_label}: {base}" if table.project_label else base


def build_comparison(selections, interval="1h", aggregation="mean",
                     overlap="intersection", min_samples=1,
                     start=None, end=None, convert_units_flag=True,
                     stratification=False) -> BuildResult:
    """
    selections: list of (TableInfo, ColumnInfo)
    overlap:    "intersection" -> only bins where every series has data
                "union"        -> keep every bin, leave gaps blank
    """
    if not selections:
        raise ValueError("nothing selected")

    frames, counts, units, cadences, sources = {}, {}, {}, {}, {}
    geometry, trust, colmap = {}, {}, {}
    dropped: list[str] = []

    for table, col in selections:
        s, unit = load_series(table, col, convert=convert_units_flag)
        if s.empty:
            # Never drop a requested series silently -- LJPC1's temperature
            # columns exist but are entirely null, and a quiet skip is how
            # that becomes an invisible hole in someone's analysis.
            dropped.append(f"{col.label} (no usable values in the source)")
            continue
        label = series_label(table, col)
        if label in frames:                     # de-duplicate across workbooks
            label = f"{table.file.rsplit('.', 1)[0]}.{label}"

        cad = native_cadence(s)
        cadences[label] = str(cad) if cad is not None else "irregular"
        units[label] = unit
        sources[label] = f"{table.file} :: {table.name} :: {col.column}"
        geometry[label] = col.geometry_label
        trust[label] = table.time_basis_note
        colmap[label] = col

        r = s.resample(interval)
        frames[label] = getattr(r, AGGREGATIONS[aggregation])()
        counts[label] = r.count()

    if not frames:
        raise ValueError("every selected series was empty after loading")

    data = pd.concat(frames, axis=1)
    cnt = pd.concat(counts, axis=1).reindex(data.index).fillna(0).astype(int)

    # A bin backed by too few raw samples is worse than no bin at all.
    if min_samples > 1:
        data = data.where(cnt >= min_samples)

    if overlap == "intersection":
        keep = data.notna().all(axis=1)
        data, cnt = data[keep], cnt[keep]
    else:
        keep = data.notna().any(axis=1)
        data, cnt = data[keep], cnt[keep]

    if start is not None:
        data, cnt = data[data.index >= start], cnt[cnt.index >= start]
    if end is not None:
        data, cnt = data[data.index <= end], cnt[cnt.index <= end]

    result = BuildResult(data, cnt, units, cadences, sources,
                         interval, aggregation, overlap, min_samples, dropped,
                         geometry, trust, colmap)

    if stratification:
        add_stratification_index(result)
    return result


def find_stratification_pair(result: BuildResult) -> tuple[str, str] | None:
    """(surface column, subsurface column) if both endpoints are present.

    SST(46254) - T(autoss) is surface minus 5 m. It is the covariate that says
    when the buoy and the pier should be expected to agree, which is why it is
    offered rather than a raw 46254-vs-buoy comparison: 46254 follows the sea
    surface and the buoy sits on the seabed, so they are not peers.
    """
    surface = subsurface = None
    for label, col in result.columns.items():
        if col.variable != "sea_water_temperature" and \
                "temperature" not in str(col.column).lower():
            continue
        if col.station == "46254" or "46254" in label:
            surface = surface or label
        elif col.station == "autoss" or "sccoos" in label.lower():
            subsurface = subsurface or label
    return (surface, subsurface) if surface and subsurface else None


def add_stratification_index(result: BuildResult) -> str | None:
    """Append SST(46254) - T(autoss) as a derived column. Returns its label."""
    pair = find_stratification_pair(result)
    if pair is None:
        return None
    surface, subsurface = pair
    # Square brackets, not parentheses: the legend formatter strips `(...)` to
    # drop unit suffixes like `(degree_Celsius)`, and it would eat the station
    # names here too.
    label = "stratification SST[46254] - T[autoss]"
    if label in result.data.columns:
        return label
    result.data[label] = result.data[surface] - result.data[subsurface]
    result.counts[label] = result.counts[[surface, subsurface]].min(axis=1)
    result.units[label] = result.units.get(surface, "degC")
    result.cadences[label] = result.cadences.get(surface, "")
    result.sources[label] = f"derived: {surface} minus {subsurface}"
    result.geometry[label] = "(0.45 m surface - 5 m earth)"
    result.time_trust[label] = "derived from the two series above"
    result.derived.append(label)
    return label


def lag_scan(data: pd.DataFrame, reference: str, interval: str,
             max_hours: float = 24.0) -> pd.DataFrame:
    """
    Cross-correlate every series against `reference` across a range of lags.

    Note for this dataset: the dominant period in La Jolla nearshore
    temperature is the ~12.4 h internal tide, so a peak at lag L is
    indistinguishable from one at L +/- 12.4 h. The `ambiguous_alt` column
    spells that alternative out rather than letting you forget it.
    """
    step = pd.Timedelta(interval).total_seconds() / 3600.0
    if step <= 0:
        raise ValueError("bad interval")
    n = int(max_hours / step)
    rows = []

    for col in data.columns:
        if col == reference:
            continue
        best_lag, best_r = 0.0, np.nan
        for k in range(-n, n + 1):
            r = data[col].corr(data[reference].shift(k))
            if pd.notna(r):
                if pd.isna(best_r) or abs(r) > abs(best_r):
                    best_lag, best_r = k * step, r
        rows.append({
            "series": col,
            "r_at_lag_0": data[col].corr(data[reference]),
            # Sign convention: negative means this series LEADS the reference
            # (its features show up earlier); positive means it lags.
            "best_lag_h": best_lag,
            "r_at_best_lag": best_r,
            "ambiguous_alt_h": best_lag - 12.42 if best_lag > 0 else best_lag + 12.42,
            "n_overlapping": int((data[col].notna() & data[reference].notna()).sum()),
        })
    return pd.DataFrame(rows)
